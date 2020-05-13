import optparse
import sys
import os
import openpyxl
import datetime
# import PyPDF2


def my_decorator(func):
    def wrap_func():
        print('******************')
        func()
        print('******************')
    return wrap_func


@my_decorator
def welcome():
    print('AIIA Resul Creator')


term = ''
session = ''
grade = ''
daysOpened = 0
daysAbsent = 0
daysPresent = daysOpened - daysAbsent
noOfPupils = 0
teacher_name = ''
resumption_date = ''


def get_term_data():
    template_year = datetime.datetime.now().year
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb.active
    session = input('Enter session: ')
    term = input('Enter term: ')
    grade = input('Enter class name: ')
    noOfPupils = input('Number of Pupils: ')
    daysOpened = input('No of Days School opened: ')
    daysAbsent = input('No of Days School absent: ')
    teacher_name = input('Teacher\'s name: ')
    resumption_date = input('Enter school resumption date: ')
    ws['G5'] = f'CLASS: {grade.upper()}'
    ws['J5'] = f'SESSION: {session.upper()}'
    ws['L5'] = f'TERM: {term.upper()}'
    ws['E41'] = teacher_name.upper()
    ws['A43'] = f'RESUMPTION DATE: {resumption_date.upper()}'
    wb.save(f'{template_year}{term.upper()}Term.xlsx')
    print('Template Created!!')
    create_new_result()


def create_new_result():
    answer = ''
    kontinue = input("Compute another result? ")
    while kontinue != 'no':
        if kontinue == 'yes':
            create_student_result()
            break
        else:
            print("Invalid answer use 'yes' or 'no'")
            create_new_result()
            break
    else:
        print('Compiling all results to one pdf')


def make_student_result():
    template_year = datetime.datetime.now().year
    wb = openpyxl.load_workbook(f'{template_year}{term.upper()}Term.xlsx')
    ws = wb.active
    student_name = ''
    student_name = input('Student name: ')
    ws['C5'] = student_name
    # dummy data
    ClassWork = 0
    HomeWork = 0
    CA = 0
    Activity = 0
    Exam = 0
    # end dummy
    mathClassWork = 0
    mathHomeWork = 0
    mathCA = 0
    mathActivity = 0
    mathExam = 0
    # Get the actual scores for the current student
    mathClassWork = input('Mathematics Class Work 10%: ')
    mathHomeWork = input('Mathematics Home Work 10%: ')
    mathActivity = input('Mathematics Activity 10%: ')
    mathCA = input('Mathematics CA 20%: ')
    mathExam = input('Mathematics Exams Work 50%: ')
    ws['F12'] = mathClassWork
    ws['G12'] = mathHomeWork
    ws['H12'] = mathCA
    ws['I12'] = mathActivity
    # Reading and Comprehensive
    rcClassWork = 0
    rcHomeWork = 0
    rcCA = 0
    rcActivity = 0
    rcExam = 0
    rcClassWork = int(input('Reading & Comp. Class Work 10%: '))
    rcHomeWork = int(input('Reading & Comp. Class Work 10%: '))
    rcCA = int(input('Reading & Comp. Class Work 10%: '))
    rcActivity = int(input('Reading & Comp. Class Work 10%: '))
    rcExam = int(input('Reading & Comp. Class Work 10%: '))
    ws['F9'] = rcClassWork
    ws['G9'] = rcHomeWork
    ws['H9'] = rcCA
    ws['I9'] = rcActivity
    ws.SaveAs(f'{student_name}.pdf')
    ws.close()


def create_student_result():
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb.active
    print('******************\nNew Student Result\n******************')
    student_name = input('Studnet name: ')
    make_student_result()
    create_new_result()


welcome()
get_term_data()
# create_student_result()
