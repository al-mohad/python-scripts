import openpyxl
get_term_data
wb = openpyxl.load_workbook('result.xlsx')
ws = wb.active

studentName = input('Enter Student name: ')
studentClass = input('Enter Student class: ')
schoolSession = input('Enter session: ')
schoolTerm = input('Enter result term: ')
noInClass = input('Enter number in Class: ')
daysSchoolOpen = input('No of Days School opened: ')
daysSchoolIsPresent = input('No of Days School present: ')
daysSchoolIsAbsent = input('No of Days School absent: ')
# get English Scores
englishTest = input('English Test 20%: ')
englishProject = input('English Project 10%: ')
englishHomeWork = input('English Home Work 10%: ')
englishClassWork = input('English Class Work 10%: ')
englishExams = input('English Exams 50%: ')

# get Maths Scores
mathsTest = input('Mathematics Test 20%: ')
mathsProject = input('Mathematics Project 10%: ')
mathsHomeWork = input('Mathematics Home Work 10%: ')
mathsClassWork = input('Mathematics Class Work 10%: ')
mathsExams = input('Mathematics Exams 50%: ')

ws['B6'] = studentName
ws['B8'] = studentClass
ws['B9'] = schoolSession
ws['B10'] = schoolTerm
ws['B11'] = noInClass
ws['J9'] = daysSchoolOpen
ws['J10'] = daysSchoolIsPresent

# English Scores
ws['D15'] = englishTest
ws['E15'] = englishProject
ws['F15'] = englishHomeWork
ws['G15'] = englishClassWork
ws['H15'] = englishExams


def get_term_data():
    print('[+]************************************************************[+]')
    return


wb.save(f'{studentName}.xlsx')
